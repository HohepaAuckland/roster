"""Hohepa Auckland Roster Management App"""
from flask import Flask, render_template_string, request, jsonify, send_file, session
import json, os, io
from collections import defaultdict
from datetime import datetime
import pandas as pd
from parser import parse_roster, compute_staff_hours, HOUSES, HOUSE_SHORT, HOUSE_FULL
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'hohepa-roster-2026'

# In-memory store (reset on restart, saved via download)
STATE = {
    'roster': {},       # house -> date -> [slots]
    'staff_map': {},    # name -> {contract, type}
    'dates': [],
    'loaded': False,
    'period': '',
}

HOUSE_COLORS = {
    'Daffodil':    '#0D3D6B', 'Christopher': '#7A2E00',
    'Hilary':      '#0A5C2E', 'Gabriel':     '#3B1F80',
    'Parzival':    '#7A4700', 'Michael':     '#7A0D5C',
    'Magnolia':    '#085A45', 'Wake':        '#1A1A2E',
}
HOUSE_LIGHT = {
    'Daffodil':    '#CDE4F5', 'Christopher': '#FAD5BB',
    'Hilary':      '#C8EDD8', 'Gabriel':     '#E0D9FF',
    'Parzival':    '#FAE0BB', 'Michael':     '#FCD5EF',
    'Magnolia':    '#C5EDE5', 'Wake':        '#D0D0E8',
}
ALERT_96 = 96.0
WARN_80  = 80.0

HTML = r"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Hohepa Auckland — Roster System</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',Arial,sans-serif;background:#F0F4F8;color:#1a1a2e;min-height:100vh}
:root{--navy:#1B3A6B;--teal:#0A5C3A;--warn:#CC6600;--danger:#CC0000;--ok:#1B5E20;--grey:#666}

/* NAV */
nav{background:var(--navy);color:#fff;padding:0 24px;display:flex;align-items:center;gap:20px;height:54px;box-shadow:0 2px 8px #0003}
nav h1{font-size:18px;font-weight:700;letter-spacing:.3px}
nav span{font-size:12px;opacity:.7;margin-left:4px}
.nav-tabs{display:flex;gap:4px;margin-left:auto}
.tab{padding:8px 16px;border:none;background:transparent;color:#ffffffaa;cursor:pointer;border-radius:6px 6px 0 0;font-size:13px;font-weight:500;transition:all .15s}
.tab.active,.tab:hover{background:#ffffff22;color:#fff}
.tab.active{background:#ffffff33;border-bottom:2px solid #fff}

/* MAIN */
main{padding:20px 24px;max-width:1400px;margin:0 auto}
.page{display:none}.page.active{display:block}

/* CARDS */
.card{background:#fff;border-radius:12px;padding:20px;box-shadow:0 1px 4px #0001;margin-bottom:16px}
.card h2{font-size:15px;font-weight:600;color:var(--navy);margin-bottom:14px;display:flex;align-items:center;gap:8px}
.card h2 .badge{font-size:11px;font-weight:500;padding:2px 10px;border-radius:20px;background:#EEF2FF;color:var(--navy)}

/* UPLOAD */
.upload-zone{border:2px dashed #B0C4DE;border-radius:10px;padding:40px;text-align:center;cursor:pointer;transition:all .2s;background:#F8FBFF}
.upload-zone:hover,.upload-zone.drag{border-color:var(--navy);background:#EEF4FF}
.upload-zone p{color:#666;font-size:14px;margin-top:8px}
.upload-zone .icon{font-size:40px}

/* BUTTONS */
.btn{padding:8px 18px;border:none;border-radius:8px;cursor:pointer;font-size:13px;font-weight:600;transition:all .15s;display:inline-flex;align-items:center;gap:6px}
.btn-primary{background:var(--navy);color:#fff}.btn-primary:hover{background:#0D2A50}
.btn-success{background:#1B5E20;color:#fff}.btn-success:hover{background:#145018}
.btn-warn{background:#CC6600;color:#fff}.btn-warn:hover{background:#A85200}
.btn-sm{padding:5px 12px;font-size:12px}
.btn-ghost{background:transparent;color:var(--navy);border:1.5px solid #B0C4DE}.btn-ghost:hover{background:#F0F4FF}

/* GRID STATS */
.stats-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:12px;margin-bottom:20px}
.stat{background:#fff;border-radius:10px;padding:14px 16px;border-left:4px solid var(--navy)}
.stat .val{font-size:26px;font-weight:700;color:var(--navy)}
.stat .lbl{font-size:11px;color:#666;margin-top:2px}
.stat.warn{border-color:var(--warn)}.stat.warn .val{color:var(--warn)}
.stat.danger{border-color:var(--danger)}.stat.danger .val{color:var(--danger)}
.stat.ok{border-color:var(--ok)}.stat.ok .val{color:var(--ok)}

/* ALERT BAR */
.alert{padding:10px 16px;border-radius:8px;font-size:13px;margin-bottom:12px;display:flex;align-items:center;gap:10px}
.alert-danger{background:#FDE8E8;color:#CC0000;border:1px solid #F5B5B5}
.alert-warn{background:#FFF3CD;color:#7A5A00;border:1px solid #FFD080}
.alert-info{background:#E8F4FD;color:#0A3D6E;border:1px solid #B0D4F0}

/* TABLE */
.tbl-wrap{overflow-x:auto;border-radius:8px;border:1px solid #E0E8F0}
table{width:100%;border-collapse:collapse;font-size:13px}
thead th{background:var(--navy);color:#fff;padding:9px 12px;text-align:left;font-weight:600;font-size:12px;white-space:nowrap}
tbody tr:nth-child(even){background:#F8FBFF}
tbody tr:hover{background:#EEF4FF}
td{padding:8px 12px;border-bottom:1px solid #E8EFF5;vertical-align:middle}
td.name-cell{font-weight:600;color:var(--navy)}
.hrs-bar{display:flex;align-items:center;gap:8px}
.bar{height:8px;border-radius:4px;min-width:4px;transition:width .3s}
.bar-ok{background:#1B5E20}.bar-warn{background:#CC6600}.bar-danger{background:#CC0000}
.hrs-num{font-weight:700;font-size:13px;min-width:38px}
.badge-house{display:inline-block;font-size:10px;font-weight:600;padding:2px 7px;border-radius:10px;margin:1px;color:#fff}
.status-ok{color:#1B5E20;font-weight:600}
.status-warn{color:#CC6600;font-weight:600}
.status-danger{color:#CC0000;font-weight:600}

/* HOUSE ROSTER GRID */
.house-select{display:flex;flex-wrap:wrap;gap:8px;margin-bottom:16px}
.house-btn{padding:7px 16px;border-radius:8px;border:2px solid transparent;cursor:pointer;font-size:13px;font-weight:600;color:#fff;transition:all .15s;opacity:.7}
.house-btn.active,.house-btn:hover{opacity:1;transform:translateY(-1px);box-shadow:0 3px 8px #0002}
.roster-grid{overflow-x:auto}
.roster-table{border-collapse:collapse;min-width:900px;width:100%}
.roster-table th{padding:8px 10px;font-size:11px;font-weight:700;text-align:center;color:#fff;white-space:nowrap}
.roster-table td{padding:6px 8px;border:1px solid #E0E8F0;vertical-align:top;min-width:90px}
.roster-table .row-label{font-size:11px;font-weight:700;color:#fff;padding:6px 10px;white-space:nowrap;min-width:120px}
.shift-cell{background:#fff;border-radius:6px;padding:5px 7px;cursor:pointer;transition:all .15s;border:1.5px solid transparent}
.shift-cell:hover{border-color:var(--navy);background:#EEF4FF}
.shift-cell .sname{font-weight:700;font-size:12px;color:#1a1a2e}
.shift-cell .stime{font-size:10px;color:#666;margin-top:1px}
.shift-cell .shrs{font-size:10px;font-weight:600;color:var(--navy)}
.shift-cell.leave{background:#FFF9E6;border-color:#FFD080}
.shift-cell.open{background:#FDE8E8;border-color:#F5B5B5}
.shift-cell .leave-tag{font-size:9px;font-weight:700;color:#CC6600;background:#FFEAA0;padding:1px 5px;border-radius:4px;margin-top:2px;display:inline-block}
.empty-cell{background:#F8F9FA;border-radius:6px;padding:5px 7px;min-height:38px;border:1.5px dashed #D0D8E0;cursor:pointer}
.empty-cell:hover{border-color:#90A8C0;background:#EEF4FF}
.weekend{background:#F5F0FF!important}

/* MODAL */
.overlay{display:none;position:fixed;inset:0;background:#0005;z-index:100;align-items:center;justify-content:center}
.overlay.open{display:flex}
.modal{background:#fff;border-radius:14px;padding:28px;max-width:480px;width:90%;box-shadow:0 8px 32px #0003;max-height:90vh;overflow-y:auto}
.modal h3{font-size:16px;font-weight:700;color:var(--navy);margin-bottom:18px}
.form-group{margin-bottom:14px}
.form-group label{display:block;font-size:12px;font-weight:600;color:#444;margin-bottom:5px}
.form-group input,.form-group select,.form-group textarea{width:100%;padding:8px 12px;border:1.5px solid #D0D8E4;border-radius:8px;font-size:13px;outline:none;transition:border .15s}
.form-group input:focus,.form-group select:focus{border-color:var(--navy)}
.form-row{display:grid;grid-template-columns:1fr 1fr;gap:12px}
.modal-actions{display:flex;gap:10px;justify-content:flex-end;margin-top:20px}

/* SEARCH */
.search-bar{position:relative;margin-bottom:16px}
.search-bar input{width:100%;padding:9px 14px 9px 38px;border:1.5px solid #D0D8E4;border-radius:8px;font-size:13px;outline:none}
.search-bar input:focus{border-color:var(--navy)}
.search-bar .icon{position:absolute;left:12px;top:50%;transform:translateY(-50%);color:#888;font-size:16px}

/* INDIVIDUAL CARD */
.staff-card{background:#fff;border-radius:10px;border:1px solid #E0E8F0;padding:16px;margin-bottom:12px}
.staff-card .card-header{display:flex;align-items:center;gap:12px;margin-bottom:12px}
.avatar{width:42px;height:42px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:16px;font-weight:700;color:#fff;flex-shrink:0}
.card-name{font-weight:700;font-size:15px;color:var(--navy)}
.card-sub{font-size:12px;color:#666;margin-top:1px}
.shift-list{display:flex;flex-direction:column;gap:4px}
.shift-row{display:flex;align-items:center;gap:8px;padding:5px 8px;border-radius:6px;font-size:12px;background:#F8FBFF}
.shift-row .date{color:#666;min-width:72px}
.shift-row .house-tag{font-weight:700;font-size:11px;padding:2px 7px;border-radius:6px;color:#fff;min-width:36px;text-align:center}
.shift-row .time{flex:1;color:#333}
.shift-row .hrs{font-weight:700;color:var(--navy);min-width:32px;text-align:right}

/* RESPONSIVE */
@media(max-width:768px){nav{flex-wrap:wrap;height:auto;padding:10px}.nav-tabs{width:100%;overflow-x:auto}.stats-grid{grid-template-columns:1fr 1fr}}

/* ADD STAFF FORM */
.add-staff-panel{background:#EEF4FF;border:1.5px solid #B0C8E8;border-radius:10px;padding:16px;margin-bottom:16px}
</style>
</head>
<body>

<nav>
  <h1>🏡 Hohepa Roster</h1>
  <span id="nav-period"></span>
  <div class="nav-tabs">
    <button class="tab active" onclick="showPage('dashboard')">📊 Dashboard</button>
    <button class="tab" onclick="showPage('houses')">🏠 Houses</button>
    <button class="tab" onclick="showPage('staff')">👤 Staff Summary</button>
    <button class="tab" onclick="showPage('individual')">📋 Individual</button>
    <button class="tab" onclick="showPage('manage')">⚙️ Manage</button>
  </div>
</nav>

<main>

<!-- DASHBOARD -->
<div id="page-dashboard" class="page active">
  <div id="upload-section">
    <div class="card">
      <h2>📂 Upload Roster Excel</h2>
      <div class="upload-zone" id="drop-zone" onclick="document.getElementById('file-input').click()"
           ondragover="ev.preventDefault();this.classList.add('drag')"
           ondragleave="this.classList.remove('drag')"
           ondrop="handleDrop(event)">
        <div class="icon">📊</div>
        <p><strong>Click to upload</strong> or drag &amp; drop your roster Excel file</p>
        <p style="font-size:12px;margin-top:6px;color:#999">Supports Poonam-format roster — Daffodil, Christopher, Hilary, Gabriel, Parzival, Michael, Magnolia, Wake sheets</p>
      </div>
      <input type="file" id="file-input" accept=".xlsx,.xls" style="display:none" onchange="uploadFile(this)">
      <div id="upload-status" style="margin-top:12px"></div>
    </div>
  </div>

  <div id="dashboard-content" style="display:none">
    <div id="alerts-section"></div>
    <div class="stats-grid" id="stats-grid"></div>
    <div class="card">
      <h2>🏠 House Summary <span class="badge" id="house-period"></span></h2>
      <div class="tbl-wrap">
        <table id="house-summary-table">
          <thead><tr>
            <th>House</th>
            <th>Budget Hrs</th>
            <th>Actual Hrs</th>
            <th>vs Budget</th>
            <th>Staff Count</th>
            <th>Avg Hrs/Staff</th>
          </tr></thead>
          <tbody id="house-summary-body"></tbody>
        </table>
      </div>
    </div>
  </div>
</div>

<!-- HOUSES -->
<div id="page-houses" class="page">
  <div class="card">
    <h2>🏠 House Rosters</h2>
    <div class="house-select" id="house-buttons"></div>
    <div id="house-roster-area">
      <p style="color:#888;font-size:13px">Upload a roster file and select a house to view.</p>
    </div>
  </div>
</div>

<!-- STAFF SUMMARY -->
<div id="page-staff" class="page">
  <div class="card">
    <h2>👤 Staff Hours Summary</h2>
    <div class="search-bar">
      <span class="icon">🔍</span>
      <input type="text" id="staff-search" placeholder="Search staff name…" oninput="filterStaff()">
    </div>
    <div id="alerts-96" style="margin-bottom:12px"></div>
    <div class="tbl-wrap">
      <table>
        <thead><tr>
          <th>Staff Name</th><th>Contract</th><th>Hours This Period</th>
          <th>DC</th><th>CH</th><th>HH</th><th>GH</th><th>PH</th><th>MiH</th><th>MaG</th><th>Wake</th>
          <th>Days</th><th>Status</th>
        </tr></thead>
        <tbody id="staff-table-body"></tbody>
      </table>
    </div>
  </div>
</div>

<!-- INDIVIDUAL -->
<div id="page-individual" class="page">
  <div class="card">
    <h2>📋 Individual Staff Roster</h2>
    <div class="search-bar">
      <span class="icon">🔍</span>
      <input type="text" id="ind-search" placeholder="Search to find a staff member…" oninput="filterIndividual()">
    </div>
    <div id="individual-cards"></div>
  </div>
</div>

<!-- MANAGE -->
<div id="page-manage" class="page">
  <div class="card">
    <h2>⚙️ Manage Staff</h2>
    <div class="add-staff-panel">
      <h3 style="font-size:14px;font-weight:600;color:var(--navy);margin-bottom:12px">➕ Add New Staff Member</h3>
      <div class="form-row">
        <div class="form-group"><label>Full Name</label><input id="new-name" placeholder="e.g. Jane Smith"></div>
        <div class="form-group"><label>Contract Hours (per fortnight)</label><input id="new-hrs" type="number" placeholder="80"></div>
      </div>
      <div class="form-row">
        <div class="form-group"><label>Contract Type</label>
          <select id="new-type">
            <option value="">Permanent</option><option value="Casual">Casual</option>
            <option value="Volunteers">Volunteer</option><option value="CP">CP</option>
          </select>
        </div>
        <div class="form-group"><label>Primary House</label>
          <select id="new-house">
            <option value="">Select house</option>
            <option>Daffodil</option><option>Christopher</option><option>Hilary</option>
            <option>Gabriel</option><option>Parzival</option><option>Michael</option>
            <option>Magnolia</option><option>Wake</option>
          </select>
        </div>
      </div>
      <button class="btn btn-primary" onclick="addStaff()">➕ Add Staff Member</button>
    </div>
    <div class="tbl-wrap">
      <table>
        <thead><tr><th>Name</th><th>Contract Hrs</th><th>Type</th><th>This Period</th><th>Status</th><th>Actions</th></tr></thead>
        <tbody id="manage-table-body"></tbody>
      </table>
    </div>
  </div>
  <div class="card">
    <h2>📥 Export</h2>
    <div style="display:flex;gap:12px;flex-wrap:wrap">
      <button class="btn btn-success" onclick="downloadExcel()">📊 Download Updated Excel</button>
      <button class="btn btn-ghost" onclick="downloadJSON()">📄 Export Data (JSON)</button>
    </div>
    <p style="font-size:12px;color:#888;margin-top:10px">Download the Excel to share house rosters with staff. JSON export preserves all edits.</p>
  </div>
</div>

</main>

<!-- EDIT SHIFT MODAL -->
<div class="overlay" id="shift-modal">
  <div class="modal">
    <h3 id="modal-title">Edit Shift</h3>
    <input type="hidden" id="m-house"><input type="hidden" id="m-date"><input type="hidden" id="m-slot">
    <div class="form-group">
      <label>Staff Name</label>
      <select id="m-name"></select>
    </div>
    <div class="form-row">
      <div class="form-group"><label>Shift Time</label>
        <select id="m-shift">
          <option value="">-- Day Off --</option>
          <option>6:00 AM - 2:00 PM</option><option>7:00 AM - 3:00 PM</option>
          <option>8:00 AM - 4:00 PM</option><option>8:00 AM - 6:00 PM</option>
          <option>9:00 AM - 5:00 PM</option><option>10:00 AM - 6:00 PM</option>
          <option>1:00 PM - 9:00 PM</option><option>2:00 PM - 10:00 PM</option>
          <option>9:30 PM - 6:30 AM</option><option>10:00 PM - 6:00 AM</option>
          <option>6:00 AM - 9:00 AM</option><option>7:00 AM - 9:00 AM</option>
        </select>
      </div>
      <div class="form-group"><label>Hours</label>
        <input type="number" id="m-hours" step="0.5" min="0" max="24" placeholder="8">
      </div>
    </div>
    <div class="form-row">
      <div class="form-group"><label>Leave Type (if applicable)</label>
        <select id="m-leave">
          <option value="">None</option>
          <option value="AL">AL — Annual Leave</option><option value="SL">SL — Sick Leave</option>
          <option value="BL">BL — Bereavement</option><option value="ML">ML — Maternity</option>
          <option value="ACC">ACC</option><option value="LWOP">LWOP</option>
          <option value="ALT">ALT — Alternative</option><option value="RDO">RDO</option>
          <option value="Training">Training</option><option value="Admin Day (HL)">Admin Day</option>
          <option value="Open Shift">Open Shift</option>
        </select>
      </div>
      <div class="form-group"><label>Replacement (if on leave)</label>
        <select id="m-replacement"></select>
      </div>
    </div>
    <div class="form-group"><label>Notes</label>
      <textarea id="m-notes" rows="2" placeholder="Any notes…" style="resize:vertical"></textarea>
    </div>
    <div class="modal-actions">
      <button class="btn btn-ghost" onclick="closeModal()">Cancel</button>
      <button class="btn btn-warn" onclick="clearSlot()">🗑 Clear Slot</button>
      <button class="btn btn-primary" onclick="saveSlot()">💾 Save</button>
    </div>
  </div>
</div>

<script>
// ── State ────────────────────────────────────────────────────────────────────
let roster = {};      // house -> date -> [{name,shift,hours,leave,replacement,notes}]
let staffMap = {};    // name -> {contract,type}
let dates = [];
let period = '';
let selectedHouse = '';

const HOUSE_COLORS = {{ house_colors | tojson }};
const HOUSE_SHORT  = {{ house_short | tojson }};
const HOUSE_ORDER  = ['Daffodil','Christopher','Hilary','Gabriel','Parzival','Michael','Magnolia','Wake'];
const BUDGETS      = {Daffodil:480,Christopher:596,Hilary:192,Gabriel:632,Parzival:438,Michael:638,Magnolia:180,Wake:608};

// ── Navigation ───────────────────────────────────────────────────────────────
function showPage(id){
  document.querySelectorAll('.page').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.getElementById('page-'+id).classList.add('active');
  event.target.classList.add('active');
  if(id==='staff') renderStaffTable();
  if(id==='houses') renderHouseButtons();
  if(id==='individual') renderIndividual();
  if(id==='manage') renderManage();
}

// ── Upload ───────────────────────────────────────────────────────────────────
function handleDrop(ev){
  ev.preventDefault();
  document.getElementById('drop-zone').classList.remove('drag');
  const file=ev.dataTransfer.files[0];
  if(file) uploadFileObj(file);
}
function uploadFile(input){ if(input.files[0]) uploadFileObj(input.files[0]); }
function uploadFileObj(file){
  const fd=new FormData(); fd.append('file',file);
  document.getElementById('upload-status').innerHTML='<div class="alert alert-info">⏳ Parsing roster…</div>';
  fetch('/upload',{method:'POST',body:fd})
    .then(r=>r.json()).then(data=>{
      if(data.error){
        document.getElementById('upload-status').innerHTML=`<div class="alert alert-danger">❌ ${data.error}</div>`;
        return;
      }
      roster=data.roster; staffMap=data.staff_map; dates=data.dates; period=data.period;
      document.getElementById('upload-status').innerHTML=`<div class="alert alert-info">✅ Loaded ${data.period}</div>`;
      document.getElementById('nav-period').textContent=period;
      document.getElementById('upload-section').style.display='none';
      document.getElementById('dashboard-content').style.display='block';
      renderDashboard();
    }).catch(e=>{ document.getElementById('upload-status').innerHTML=`<div class="alert alert-danger">❌ Error: ${e}</div>`; });
}

// ── Dashboard ────────────────────────────────────────────────────────────────
function renderDashboard(){
  const hrs96=[],hrsLow=[],hrsOk=[];
  const staffHours=computeStaffHours();
  Object.entries(staffHours).forEach(([name,d])=>{
    const ch=getContract(name);
    if(d.total>=96) hrs96.push({name,total:d.total});
    else if(ch&&d.total<ch-2) hrsLow.push({name,total:d.total,contract:ch});
    else if(ch) hrsOk.push(name);
  });

  // Alerts
  let alertsHtml='';
  if(hrs96.length){
    alertsHtml+=`<div class="alert alert-danger">🚨 <strong>${hrs96.length} staff at or above 96h threshold:</strong> ${hrs96.map(x=>`${x.name} (${x.total.toFixed(1)}h)`).join(', ')}</div>`;
  }
  if(hrsLow.length){
    alertsHtml+=`<div class="alert alert-warn">⚠️ <strong>${hrsLow.length} staff under contracted hours:</strong> ${hrsLow.map(x=>`${x.name} (${x.total.toFixed(1)}/${x.contract}h)`).slice(0,6).join(', ')}${hrsLow.length>6?'…':''}</div>`;
  }
  document.getElementById('alerts-section').innerHTML=alertsHtml;

  // Stats
  const totalH=Object.values(staffHours).reduce((s,d)=>s+d.total,0);
  const activeStaff=Object.keys(staffHours).filter(n=>staffHours[n].total>0).length;
  document.getElementById('stats-grid').innerHTML=`
    <div class="stat"><div class="val">${activeStaff}</div><div class="lbl">Active Staff</div></div>
    <div class="stat"><div class="val">${totalH.toFixed(0)}h</div><div class="lbl">Total Hours Rostered</div></div>
    <div class="stat danger"><div class="val">${hrs96.length}</div><div class="lbl">⚠️ At/Over 96h Alert</div></div>
    <div class="stat warn"><div class="val">${hrsLow.length}</div><div class="lbl">Under Target Hours</div></div>
    <div class="stat ok"><div class="val">${hrsOk.length}</div><div class="lbl">On Target (80h)</div></div>
    <div class="stat"><div class="val">${dates.length}</div><div class="lbl">Days in Period</div></div>
  `;

  document.getElementById('house-period').textContent=period;

  // House summary table
  const houseRows=HOUSE_ORDER.map(house=>{
    const short=HOUSE_SHORT[house]||house;
    const budget=BUDGETS[house]||0;
    let actual=0; const staffSet=new Set();
    dates.forEach(ds=>{
      (roster[house]&&roster[house][ds]||[]).forEach(slot=>{
        if(slot.name&&slot.name!=='Open Shift'){actual+=slot.hours||0; staffSet.add(slot.name);}
      });
    });
    const diff=actual-budget;
    const diffStr=diff===0?'✓ On budget':(diff>0?`+${diff.toFixed(0)}h over`:`${diff.toFixed(0)}h under`);
    const diffCls=diff>5?'status-warn':diff<-20?'status-warn':'status-ok';
    const avg=staffSet.size?actual/staffSet.size:0;
    const hcolor=HOUSE_COLORS[house]||'#333';
    return `<tr>
      <td><span style="display:inline-flex;align-items:center;gap:8px">
        <span style="width:12px;height:12px;border-radius:3px;background:${hcolor};display:inline-block"></span>
        <strong>${house}</strong></span></td>
      <td>${budget}h</td>
      <td><strong>${actual.toFixed(0)}h</strong></td>
      <td class="${diffCls}">${diffStr}</td>
      <td>${staffSet.size} staff</td>
      <td>${avg.toFixed(1)}h</td>
    </tr>`;
  });
  document.getElementById('house-summary-body').innerHTML=houseRows.join('');
}

// ── Houses ───────────────────────────────────────────────────────────────────
function renderHouseButtons(){
  const btns=HOUSE_ORDER.map(h=>`
    <button class="house-btn${selectedHouse===h?' active':''}"
      style="background:${HOUSE_COLORS[h]||'#333'}"
      onclick="selectHouse('${h}')">
      ${h}
    </button>`).join('');
  document.getElementById('house-buttons').innerHTML=btns;
  if(selectedHouse) renderHouseRoster(selectedHouse);
}
function selectHouse(h){ selectedHouse=h; renderHouseButtons(); renderHouseRoster(h); }

function renderHouseRoster(house){
  const hcolor=HOUSE_COLORS[house]||'#1B3A6B';
  const houseData=roster[house]||{};
  const sortedDates=dates.slice().sort();

  // Build rows: group slots by shift category
  const slotsByDate={};
  sortedDates.forEach(ds=>{ slotsByDate[ds]=houseData[ds]||[]; });

  // Find max slots per date
  let maxSlots=1;
  sortedDates.forEach(ds=>{ if(slotsByDate[ds].length>maxSlots) maxSlots=slotsByDate[ds].length; });

  // Date headers
  const headerCells=sortedDates.map(ds=>{
    const dt=new Date(ds+'T12:00:00');
    const isWe=dt.getDay()===0||dt.getDay()===6;
    const bg=isWe?'#3B1F80':hcolor;
    return `<th style="background:${bg};font-size:11px;text-align:center;padding:6px 8px">
      ${dt.toLocaleDateString('en-NZ',{weekday:'short'})}<br>
      <span style="font-size:13px;font-weight:700">${dt.toLocaleDateString('en-NZ',{day:'numeric',month:'short'})}</span>
    </th>`;
  }).join('');

  // Slot rows
  let bodyRows='';
  for(let slot=0;slot<maxSlots;slot++){
    const cells=sortedDates.map(ds=>{
      const slotData=(slotsByDate[ds]||[])[slot];
      if(slotData&&slotData.name){
        const isLeave=!!slotData.leave;
        const isOpen=slotData.name==='Open Shift';
        const cls=isOpen?'open':isLeave?'leave':'';
        const leaveTag=slotData.leave?`<div class="leave-tag">${slotData.leave}</div>`:'';
        const replTag=slotData.replacement?`<div style="font-size:10px;color:#1B5E20;font-weight:600">↳ ${slotData.replacement}</div>`:'';
        return `<td><div class="shift-cell ${cls}" onclick="openModal('${house}','${ds}',${slot})">
          <div class="sname">${slotData.name}</div>
          <div class="stime">${slotData.shift||''}</div>
          <div class="shrs">${slotData.hours?slotData.hours.toFixed(1)+'h':''}</div>
          ${leaveTag}${replTag}
        </div></td>`;
      } else {
        return `<td class="${isWeekend(ds)?'weekend':''}"><div class="empty-cell" onclick="openModal('${house}','${ds}',${slot})">
          <span style="font-size:10px;color:#aaa">+ add</span>
        </div></td>`;
      }
    }).join('');
    bodyRows+=`<tr>
      <td class="row-label" style="background:${hcolor}">${slotToLabel(slot)}</td>
      ${cells}
    </tr>`;
  }

  // Extra row to add slot
  const addCells=sortedDates.map(ds=>`<td><div class="empty-cell" onclick="openModal('${house}','${ds}',${maxSlots})" style="min-height:28px"><span style="font-size:10px;color:#aaa">+ slot</span></div></td>`).join('');
  bodyRows+=`<tr><td class="row-label" style="background:${hcolor};font-size:10px">Add Staff</td>${addCells}</tr>`;

  // Day total row
  const totCells=sortedDates.map(ds=>{
    const tot=(slotsByDate[ds]||[]).reduce((s,x)=>s+(x.hours||0),0);
    const staff=(slotsByDate[ds]||[]).filter(x=>x.name&&x.name!=='Open Shift').length;
    return `<td style="text-align:center;font-weight:700;font-size:12px;background:#EEF4FF">${tot.toFixed(0)}h<br><span style="font-size:10px;color:#666">${staff} staff</span></td>`;
  }).join('');
  bodyRows+=`<tr><td class="row-label" style="background:${hcolor}">Daily Total</td>${totCells}</tr>`;

  document.getElementById('house-roster-area').innerHTML=`
    <div class="roster-grid">
      <table class="roster-table">
        <thead><tr>
          <th style="background:${hcolor};min-width:120px">Slot</th>
          ${headerCells}
        </tr></thead>
        <tbody>${bodyRows}</tbody>
      </table>
    </div>`;
}

function slotToLabel(i){
  const labels=['Admin / HL','Morning Shift','Morning Shift 2','Afternoon Shift','Afternoon Shift 2','Night / Wake','Extra Slot'];
  return labels[i]||`Slot ${i+1}`;
}
function isWeekend(ds){ const d=new Date(ds+'T12:00:00'); return d.getDay()===0||d.getDay()===6; }

// ── Staff Table ──────────────────────────────────────────────────────────────
function computeStaffHours(){
  const result={};
  Object.entries(roster).forEach(([house,dateData])=>{
    const short=HOUSE_SHORT[house]||house;
    Object.entries(dateData).forEach(([ds,slots])=>{
      slots.forEach(slot=>{
        if(!slot.name||slot.name==='Open Shift') return;
        if(!result[slot.name]) result[slot.name]={total:0,houses:{},shifts:[],days:new Set()};
        result[slot.name].total+=(slot.hours||0);
        result[slot.name].houses[short]=(result[slot.name].houses[short]||0)+(slot.hours||0);
        result[slot.name].shifts.push({date:ds,house:short,shift:slot.shift,hours:slot.hours,leave:slot.leave});
        result[slot.name].days.add(ds);
      });
    });
  });
  // Convert Set to count
  Object.values(result).forEach(d=>{ d.dayCount=d.days.size; });
  return result;
}

function getContract(name){
  const s=staffMap[name];
  if(!s) return null;
  const v=parseFloat(s.contract);
  return isNaN(v)?null:v;
}

function renderStaffTable(filter=''){
  const staffHours=computeStaffHours();
  const SHORT_HOUSES=['DC','CH','HH','GH','PH','MiH','MaG','Wake'];
  const q=(filter||document.getElementById('staff-search').value||'').toLowerCase();

  // 96h alerts
  const over96=Object.entries(staffHours).filter(([n,d])=>d.total>=96).map(([n,d])=>n);
  document.getElementById('alerts-96').innerHTML=over96.length
    ? `<div class="alert alert-danger">🚨 <strong>96h Threshold Alert:</strong> ${over96.join(', ')}</div>` : '';

  const rows=Object.entries(staffHours)
    .filter(([name])=>!q||name.toLowerCase().includes(q))
    .sort((a,b)=>b[1].total-a[1].total)
    .map(([name,d])=>{
      const ch=getContract(name);
      const total=d.total;
      const pct=ch?Math.min(100,total/ch*100):50;
      let barCls='bar-ok',statusTxt='',statusCls='status-ok';
      if(total>=96){ barCls='bar-danger';statusTxt='🚨 Over 96h';statusCls='status-danger'; }
      else if(ch&&total>=ch){ barCls='bar-warn';statusTxt='▲ Over target';statusCls='status-warn'; }
      else if(ch&&total<ch-2){ barCls='bar-warn';statusTxt='▼ Under target';statusCls='status-warn'; }
      else if(ch){ statusTxt='✓ On target';statusCls='status-ok'; }
      else { statusTxt='Casual/Vol'; statusCls=''; }

      const houseCells=SHORT_HOUSES.map(h=>{
        const v=d.houses[h]||0;
        const houseKey=Object.keys(HOUSE_SHORT).find(k=>HOUSE_SHORT[k]===h)||h;
        const bg=v>0?HOUSE_COLORS[houseKey]+'22':'transparent';
        const col=v>0?HOUSE_COLORS[houseKey]:'#ccc';
        return `<td style="text-align:center;background:${bg}">
          ${v>0?`<span style="color:${col};font-weight:700;font-size:12px">${v.toFixed(0)}h</span>`:''}
        </td>`;
      }).join('');

      return `<tr>
        <td class="name-cell">${name}</td>
        <td style="text-align:center">${ch?ch+'h':'—'}</td>
        <td>
          <div class="hrs-bar">
            <div class="bar ${barCls}" style="width:${Math.max(4,pct*1.2)}px"></div>
            <span class="hrs-num">${total.toFixed(1)}h</span>
          </div>
        </td>
        ${houseCells}
        <td style="text-align:center">${d.dayCount}</td>
        <td class="${statusCls}">${statusTxt}</td>
      </tr>`;
    });

  document.getElementById('staff-table-body').innerHTML=rows.join('')||
    '<tr><td colspan="13" style="text-align:center;color:#888;padding:20px">No staff found</td></tr>';
}
function filterStaff(){ renderStaffTable(document.getElementById('staff-search').value); }

// ── Individual ───────────────────────────────────────────────────────────────
function renderIndividual(filter=''){
  const staffHours=computeStaffHours();
  const q=(filter||document.getElementById('ind-search').value||'').toLowerCase();
  const names=Object.keys(staffHours).filter(n=>!q||n.toLowerCase().includes(q)).sort();

  const cards=names.map(name=>{
    const d=staffHours[name];
    const ch=getContract(name);
    const total=d.total;
    let statusTxt='',statusBg='#E8F5E9',statusColor='#1B5E20';
    if(total>=96){statusTxt='🚨 '+total.toFixed(1)+'h — OVER 96h ALERT';statusBg='#FDE8E8';statusColor='#CC0000';}
    else if(ch&&total>=ch){statusTxt='▲ '+total.toFixed(1)+'h / '+ch+'h (over)';statusBg='#FFF3CD';statusColor='#996600';}
    else if(ch&&total<ch-2){statusTxt='▼ '+total.toFixed(1)+'h / '+ch+'h (under)';statusBg='#FFF3CD';statusColor='#996600';}
    else{statusTxt=(ch?total.toFixed(1)+'h / '+ch+'h ✓':total.toFixed(1)+'h — Casual');}

    const initials=name.split(' ').map(p=>p[0]).join('').toUpperCase().slice(0,2);
    const avatarColor='#'+Math.abs(name.split('').reduce((h,c)=>((h<<5)-h)+c.charCodeAt(0),0)&0xFFFFFF).toString(16).padStart(6,'0').slice(0,6);

    const shiftRows=d.shifts.sort((a,b)=>a.date.localeCompare(b.date)).map(s=>{
      const dt=new Date(s.date+'T12:00:00');
      const houseKey=Object.keys(HOUSE_SHORT).find(k=>HOUSE_SHORT[k]===s.house)||s.house;
      const hcol=HOUSE_COLORS[houseKey]||'#333';
      return `<div class="shift-row">
        <span class="date">${dt.toLocaleDateString('en-NZ',{weekday:'short',day:'numeric',month:'short'})}</span>
        <span class="house-tag" style="background:${hcol}">${s.house}</span>
        <span class="time">${s.shift||'—'}</span>
        <span class="hrs">${s.hours?s.hours.toFixed(1)+'h':''}</span>
        ${s.leave?`<span style="font-size:10px;font-weight:700;color:#CC6600;background:#FFF0C0;padding:1px 5px;border-radius:4px">${s.leave}</span>`:''}
      </div>`;
    }).join('');

    const houseBreakdown=Object.entries(d.houses).map(([h,hrs])=>{
      const houseKey=Object.keys(HOUSE_SHORT).find(k=>HOUSE_SHORT[k]===h)||h;
      const hcol=HOUSE_COLORS[houseKey]||'#333';
      return `<span class="badge-house" style="background:${hcol}">${h}: ${hrs.toFixed(0)}h</span>`;
    }).join('');

    return `<div class="staff-card">
      <div class="card-header">
        <div class="avatar" style="background:${avatarColor}">${initials}</div>
        <div>
          <div class="card-name">${name}</div>
          <div class="card-sub">${ch?'Contract: '+ch+'h per fortnight':'Casual / Volunteer'}</div>
          <div style="margin-top:4px">${houseBreakdown}</div>
        </div>
        <div style="margin-left:auto;text-align:right">
          <div style="font-size:12px;font-weight:700;color:${statusColor};background:${statusBg};padding:5px 12px;border-radius:8px">${statusTxt}</div>
          <div style="font-size:11px;color:#888;margin-top:4px">${d.dayCount} days rostered</div>
        </div>
      </div>
      <div class="shift-list">${shiftRows||'<div style="color:#aaa;font-size:12px;padding:8px">No shifts this period</div>'}</div>
    </div>`;
  });

  document.getElementById('individual-cards').innerHTML=cards.join('')||
    '<p style="color:#888;font-size:13px">No staff found. Upload a roster to see individual schedules.</p>';
}
function filterIndividual(){ renderIndividual(document.getElementById('ind-search').value); }

// ── Manage ───────────────────────────────────────────────────────────────────
function renderManage(){
  const staffHours=computeStaffHours();
  const allNames=new Set([...Object.keys(staffMap),...Object.keys(staffHours)]);
  const rows=[...allNames].sort().map(name=>{
    const sm=staffMap[name]||{};
    const ch=getContract(name);
    const total=(staffHours[name]||{total:0}).total;
    let statusTxt='',statusCls='';
    if(total>=96){statusTxt='🚨 Over 96h';statusCls='status-danger';}
    else if(ch&&total>=ch){statusTxt='▲ Over';statusCls='status-warn';}
    else if(ch&&total<ch-2){statusTxt='▼ Under';statusCls='status-warn';}
    else if(ch){statusTxt='✓ OK';statusCls='status-ok';}
    else{statusTxt='Casual';}
    return `<tr>
      <td class="name-cell">${name}</td>
      <td style="text-align:center">${ch?ch+'h':(sm.contract||'—')}</td>
      <td>${sm.type||'Permanent'}</td>
      <td style="text-align:center;font-weight:700">${total.toFixed(1)}h</td>
      <td class="${statusCls}">${statusTxt}</td>
      <td><button class="btn btn-ghost btn-sm" onclick="editStaff('${name}')">✏️ Edit</button></td>
    </tr>`;
  });
  document.getElementById('manage-table-body').innerHTML=rows.join('');
}

function addStaff(){
  const name=document.getElementById('new-name').value.trim();
  const hrs=document.getElementById('new-hrs').value;
  const type=document.getElementById('new-type').value;
  if(!name){ alert('Please enter a staff name'); return; }
  staffMap[name]={contract:parseFloat(hrs)||80,type:type};
  document.getElementById('new-name').value='';
  document.getElementById('new-hrs').value='';
  renderManage();
  alert(`✅ ${name} added to staff list`);
}

function editStaff(name){
  const newName=prompt('Edit name:',name);
  if(!newName) return;
  const newHrs=prompt('Contract hours:',staffMap[name]?.contract||80);
  if(newName!==name){
    staffMap[newName]=staffMap[name]||{};
    delete staffMap[name];
  }
  if(staffMap[newName]) staffMap[newName].contract=parseFloat(newHrs)||80;
  renderManage();
}

// ── Modal ────────────────────────────────────────────────────────────────────
function openModal(house,date,slot){
  document.getElementById('shift-modal').classList.add('open');
  document.getElementById('m-house').value=house;
  document.getElementById('m-date').value=date;
  document.getElementById('m-slot').value=slot;
  const dt=new Date(date+'T12:00:00');
  document.getElementById('modal-title').textContent=
    `${house} — ${dt.toLocaleDateString('en-NZ',{weekday:'long',day:'numeric',month:'short'})} — Slot ${parseInt(slot)+1}`;

  const existing=(roster[house]&&roster[house][date]&&roster[house][date][slot])||{};

  // Populate staff dropdowns
  const allStaff=[...new Set([...Object.keys(staffMap),...Object.keys(computeStaffHours())])].sort();
  const opts='<option value="">-- Empty / Day Off --</option>'+
    allStaff.map(n=>`<option value="${n}"${existing.name===n?' selected':''}>${n}</option>`).join('');
  document.getElementById('m-name').innerHTML=opts;
  const replOpts='<option value="">None</option>'+
    allStaff.map(n=>`<option value="${n}"${existing.replacement===n?' selected':''}>${n}</option>`).join('');
  document.getElementById('m-replacement').innerHTML=replOpts;

  document.getElementById('m-shift').value=existing.shift||'';
  document.getElementById('m-hours').value=existing.hours||'';
  document.getElementById('m-leave').value=existing.leave||'';
  document.getElementById('m-notes').value=existing.notes||'';
}

function closeModal(){ document.getElementById('shift-modal').classList.remove('open'); }

function saveSlot(){
  const house=document.getElementById('m-house').value;
  const date=document.getElementById('m-date').value;
  const slot=parseInt(document.getElementById('m-slot').value);
  const name=document.getElementById('m-name').value;
  const shift=document.getElementById('m-shift').value;
  const hours=parseFloat(document.getElementById('m-hours').value)||0;
  const leave=document.getElementById('m-leave').value;
  const replacement=document.getElementById('m-replacement').value;
  const notes=document.getElementById('m-notes').value;

  if(!roster[house]) roster[house]={};
  if(!roster[house][date]) roster[house][date]=[];
  while(roster[house][date].length<=slot) roster[house][date].push({name:'',shift:'',hours:0,leave:'',replacement:'',notes:''});
  roster[house][date][slot]={name,shift,hours,leave,replacement,notes};

  closeModal();
  if(selectedHouse===house) renderHouseRoster(house);
  renderDashboard();
}

function clearSlot(){
  const house=document.getElementById('m-house').value;
  const date=document.getElementById('m-date').value;
  const slot=parseInt(document.getElementById('m-slot').value);
  if(roster[house]&&roster[house][date]&&roster[house][date][slot]){
    roster[house][date][slot]={name:'',shift:'',hours:0,leave:'',replacement:'',notes:''};
  }
  closeModal();
  if(selectedHouse===house) renderHouseRoster(house);
}

// Close modal on overlay click
document.getElementById('shift-modal').addEventListener('click',function(e){
  if(e.target===this) closeModal();
});

// ── Export ───────────────────────────────────────────────────────────────────
function downloadExcel(){
  fetch('/export-excel',{
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body:JSON.stringify({roster,staff_map:staffMap,dates,period})
  }).then(r=>r.blob()).then(blob=>{
    const a=document.createElement('a');
    a.href=URL.createObjectURL(blob);
    a.download=`Hohepa_Roster_${new Date().toISOString().slice(0,10)}.xlsx`;
    a.click();
  });
}
function downloadJSON(){
  const blob=new Blob([JSON.stringify({roster,staff_map:staffMap,dates,period},null,2)],{type:'application/json'});
  const a=document.createElement('a');
  a.href=URL.createObjectURL(blob); a.download='roster_data.json'; a.click();
}
</script>
</body>
</html>
""".replace('{{ house_colors | tojson }}', json.dumps(HOUSE_COLORS)).replace('{{ house_short | tojson }}', json.dumps(HOUSE_SHORT))

@app.route('/')
def index():
    return render_template_string(HTML)

@app.route('/upload', methods=['POST'])
def upload():
    try:
        f = request.files.get('file')
        if not f:
            return jsonify({'error': 'No file uploaded'})
        
        buf = io.BytesIO(f.read())
        roster_data, staff_map = parse_roster(buf)
        
        # Get all dates
        all_dates = set()
        for house_data in roster_data.values():
            all_dates.update(house_data.keys())
        dates = sorted(all_dates)
        
        period = ''
        if dates:
            d0 = datetime.strptime(dates[0], '%Y-%m-%d')
            d1 = datetime.strptime(dates[-1], '%Y-%m-%d')
            period = f"{d0.strftime('%d %b')} – {d1.strftime('%d %b %Y')}"
        
        # Store in STATE
        STATE['roster']   = roster_data
        STATE['staff_map']= staff_map
        STATE['dates']    = dates
        STATE['period']   = period
        STATE['loaded']   = True
        
        return jsonify({
            'roster':    roster_data,
            'staff_map': {k: {'contract': v['contract'], 'type': v['type']} for k,v in staff_map.items()},
            'dates':     dates,
            'period':    period,
        })
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()})

@app.route('/export-excel', methods=['POST'])
def export_excel():
    data = request.get_json()
    roster   = data.get('roster', {})
    staff_map= data.get('staff_map', {})
    dates    = data.get('dates', [])
    period   = data.get('period', '')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def fl(h): return PatternFill('solid', fgColor=h)
    def fn(bold=False, color='FF000000', size=10):
        return Font(bold=bold, color=color, size=size, name='Arial')
    def bd():
        s=Side(style='thin',color='FFBBBBBB')
        return Border(left=s,right=s,top=s,bottom=s)
    def ctr(w=False): return Alignment(horizontal='center',vertical='center',wrap_text=w)
    def lft(): return Alignment(horizontal='left',vertical='center')

    HC = {k: v.replace('#','FF') for k,v in HOUSE_COLORS.items()}

    # Staff Summary sheet
    ws = wb.create_sheet('Staff Summary')
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    for ci in range(3, 12): ws.column_dimensions[get_column_letter(ci)].width = 10
    ws.column_dimensions[get_column_letter(12)].width = 14

    ws.merge_cells('A1:L1')
    ws['A1'] = f'STAFF SUMMARY — {period}'
    ws['A1'].font = fn(True,'FFFFFFFF',14)
    ws['A1'].fill = fl('FF1B3A6B'); ws['A1'].alignment = ctr()
    ws.row_dimensions[1].height = 30

    hdrs = ['Staff Name','Contract','Total Hrs','DC','CH','HH','GH','PH','MiH','MaG','Wake','Status']
    for ci,h in enumerate(hdrs,1):
        c=ws.cell(3,ci,h); c.font=fn(True,'FFFFFFFF',10)
        c.fill=fl('FF1B3A6B'); c.alignment=ctr(); c.border=bd()
    ws.row_dimensions[3].height=18

    # Compute staff hours
    staff_hrs = defaultdict(lambda:{'total':0,'houses':defaultdict(float),'days':set()})
    for house, date_data in roster.items():
        short = HOUSE_SHORT.get(house, house)
        for ds, slots in date_data.items():
            for slot in slots:
                name = slot.get('name','')
                if not name or name=='Open Shift': continue
                hrs = slot.get('hours',0) or 0
                staff_hrs[name]['total'] += hrs
                staff_hrs[name]['houses'][short] += hrs
                staff_hrs[name]['days'].add(ds)

    SHORT_H = ['DC','CH','HH','GH','PH','MiH','MaG','Wake']
    for ri,(name,d) in enumerate(sorted(staff_hrs.items()),4):
        alt = ri%2==0; bg='FFF5FAFF' if alt else 'FFFFFFFF'
        sm = staff_map.get(name,{}); ch=sm.get('contract') if sm else None
        try: ch=float(ch)
        except: ch=None
        total=d['total']
        status = '🚨 OVER 96h' if total>=96 else ('▲ Over' if ch and total>ch else ('▼ Under' if ch and total<ch-2 else '✓ OK'))
        sfg = 'FFCC0000' if total>=96 else ('FF996600' if ch and abs(total-ch)>2 else 'FF1B5E20')
        vals=[name,f'{ch:.0f}h' if ch else '—',round(total,1)]
        vals+=[round(d['houses'].get(h,0),1) or '' for h in SHORT_H]
        vals+=[status]
        for ci,v in enumerate(vals,1):
            c=ws.cell(ri,ci,v); c.fill=fl(bg); c.border=bd()
            c.font=fn(bold=(ci==1 or ci==12),color=sfg if ci==12 else 'FF000000',size=10)
            c.alignment=ctr() if ci!=1 else lft()
        ws.row_dimensions[ri].height=15

    # House sheets
    for house in HOUSES:
        house_data = roster.get(house,{})
        if not house_data: continue
        ws2 = wb.create_sheet(house[:12])
        hcolor = HC.get(house,'FF1B3A6B')
        house_dates = sorted(house_data.keys())

        ws2.column_dimensions['A'].width = 14
        for di in range(len(house_dates)):
            ws2.column_dimensions[get_column_letter(di+2)].width = 14

        ws2.merge_cells(f'A1:{get_column_letter(len(house_dates)+1)}1')
        ws2.cell(1,1,f'{house.upper()} — {period}').font=fn(True,'FFFFFFFF',13)
        ws2.cell(1,1).fill=fl(hcolor); ws2.cell(1,1).alignment=ctr()
        ws2.row_dimensions[1].height=28

        # Day headers
        ws2.cell(2,1,'Slot').font=fn(True,'FFFFFFFF',10)
        ws2.cell(2,1).fill=fl(hcolor); ws2.cell(2,1).border=bd(); ws2.cell(2,1).alignment=ctr()
        for di,ds in enumerate(house_dates):
            dt=datetime.strptime(ds,'%Y-%m-%d')
            col=di+2
            c=ws2.cell(2,col,dt.strftime('%a %d %b'))
            c.font=fn(True,'FFFFFFFF',10); c.fill=fl(hcolor)
            c.alignment=ctr(); c.border=bd()
        ws2.row_dimensions[2].height=20

        max_slots=max((len(v) for v in house_data.values()),default=0)
        for slot in range(max_slots):
            row=slot+3
            ws2.cell(row,1,f'Staff {slot+1}').font=fn(True,'FFFFFFFF',9)
            ws2.cell(row,1).fill=fl(hcolor); ws2.cell(row,1).border=bd(); ws2.cell(row,1).alignment=ctr()
            ws2.row_dimensions[row].height=18
            for di,ds in enumerate(house_dates):
                col=di+2
                slots=house_data.get(ds,[])
                sdata=slots[slot] if slot<len(slots) else {}
                name=sdata.get('name','') or ''
                shift=sdata.get('shift','') or ''
                hrs=sdata.get('hours',0) or 0
                leave=sdata.get('leave','') or ''
                bg='FFFFF3CD' if leave else ('FFFFFFFF' if name else 'FFF5F5F5')
                c=ws2.cell(row,col,f'{name}\n{shift}\n{hrs:.0f}h{" ["+leave+"]" if leave else ""}' if name else '')
                c.font=fn(bold=bool(name),size=9,color='FF1A1A3E')
                c.fill=fl(bg); c.alignment=ctr(w=True); c.border=bd()

        # Totals row
        tot_row=max_slots+3
        ws2.cell(tot_row,1,'Total').font=fn(True,'FFFFFFFF',10)
        ws2.cell(tot_row,1).fill=fl(hcolor); ws2.cell(tot_row,1).border=bd(); ws2.cell(tot_row,1).alignment=ctr()
        ws2.row_dimensions[tot_row].height=20
        for di,ds in enumerate(house_dates):
            col=di+2
            tot=sum(s.get('hours',0) or 0 for s in house_data.get(ds,[]))
            st=sum(1 for s in house_data.get(ds,[]) if s.get('name') and s['name']!='Open Shift')
            c=ws2.cell(tot_row,col,f'{tot:.0f}h / {st} staff')
            c.font=fn(True,'FFFFFFFF',10); c.fill=fl(hcolor)
            c.alignment=ctr(); c.border=bd()

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'Hohepa_Roster_{datetime.now().strftime("%Y%m%d")}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5050))
    print(f'Starting Hohepa Roster System on port {port}')
    app.run(debug=False, port=port, host='0.0.0.0')
